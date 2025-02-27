import os

import pandas as pd
from openpyxl import load_workbook


def excel_to_parquet(excel_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    wb = load_workbook(filename=excel_path, read_only=True)

    # loop through worksheets
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")

        df = pd.read_excel(excel_path, sheet_name=sheet_name)

        parquet_file_path = os.path.join(output_dir, f"{sheet_name}.parquet")

        df.to_parquet(parquet_file_path, index=False)
        print(f"Exported: {parquet_file_path}")

    print("Excel to parquet conversion completed")


excel_path = "BJGP.2022.0235_suppl_Appendix_1.xlsx"
output_dir = "pq"

# Tables pulled from the PDF
excel_path2 = "BJGP.2022.0235_suppl_Appendix_2.xlsx"


def main():
    excel_to_parquet(excel_path, output_dir)
    excel_to_parquet(excel_path2, output_dir)


if __name__ == "__main__":
    main()
